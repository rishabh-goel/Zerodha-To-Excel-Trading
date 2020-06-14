import json
from kiteconnect import KiteConnect, KiteTicker
import os
import pythoncom
import xlwings as xw
import re
from openpyxl import load_workbook, Workbook
import pandas as pd
from kiteconnect import exceptions as KiteException
from tempfile import gettempdir
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
import urllib.parse as urlparse
from urllib.parse import parse_qs

kws = None
u = None
s = None
data = None
loop = None
instrument_tokens = []


def on_ticks(ws, ticks):
    # Callback to receive ticks.
    print("Ticks= {}".format(ticks))


def on_connect(ws, response):
    # Callback on successful connect.

    global instrument_tokens
    print("Inside onconnect")
    filename = os.path.join(os.path.dirname(__file__), "Zerodha Instruments.xlsx")

    pythoncom.CoInitialize()
    app = xw.App(visible=False)
    wb = app.books.open(filename)
    sheet = wb.sheets['Watchlist']
    df = pd.read_excel(filename, sheet_name="Watchlist")
    print("DF length: {}".format(len(df)))
    for row in range(2, len(df)+2):
        instrument_tokens.append(int(sheet.range('A'+str(row)).value))

    ws.subscribe(instrument_tokens)

    ws.set_mode(ws.MODE_LTP, instrument_tokens)

    wb.save(filename)
    # wb.close()
    app.quit()


def on_close(ws, code, reason):
    # On connection close stop the event loop.
    # Reconnection will not happen after executing `ws.stop()`
    # ws.stop()
    pass


def get_instruments():
    global u
    filename = os.path.join(os.path.dirname(__file__), 'Zerodha Instruments.xlsx')
    instruments = u.instruments()
    ordered_list = ["instrument_token", "exchange_token", "tradingsymbol", "name", "last_price", "expiry", "strike",
                    "tick_size",
                    "lot_size", "instrument_type", "segment", "exchange"]

    wb = Workbook(filename)
    ws = wb.create_sheet("New Sheet")  # or leave it blank, default name is "Sheet 1"

    ws.append(ordered_list)

    for product in instruments:
        # create a `generator` yield product `value`
        # use the fieldnames in desired order as `key`
        values = (product[k] for k in ordered_list)
        # append the `generator values`
        ws.append(values)

    wb.save(filename)
    wb.close()

    trading_symbol = ["ACC", "ADANIENT", "ADANIPORTS", "AMARAJABAT", "AMBUJACEM", "APOLLOHOSP", "APOLLOTYRE", "ASHOKLEY",
                      "ASIANPAINT", "AUROPHARMA", "AXISBANK", "BAJAJ-AUTO", "BAJAJFINSV", "BALKRISIND", "BANDHANBNK",
                      "BATAINDIA", "BEL", "BERGEPAINT", "BHARATFORG", "BHARTIARTL", "BIOCON", "BOSCHLTD", "BPCL",
                      "BRITANNIA", "CADILAHC", "CANBK", "CENTURYTEX", "CESC", "CHOLAFIN", "CIPLA", "COALINDIA", "COLPAL",
                      "CONCOR", "CUMMINSIND", "DABUR", "DIVISLAB", "DLF", "DRREDDY", "EICHERMOT", "EQUITAS", "ESCORTS",
                      "EXIDEIND", "GAIL", "GLENMARK", "GODREJCP", "GRASIM", "HAVELLS", "HCLTECH", "RAJABAH",
                      "CESC20MAYFUT", "BAJFINANCE20MAYFUT", "CRUDEOIL20MAYFUT", "GOLDM20MAYFUT", "CENTURYTEX20MAYFUT",
                      "IDEA20MAYFUT", "HDFCBANK", "HDFCLIFE", "HEROMOTOCO", "HINDALCO", "HINDPETRO", "HINDUNILVR", "IBULHSGFIN", "IDEA",
                      "ICICIBANK", "ICICIPRULI", "IGL", "INDIGO", "INDUSINDBK", "INFRATEL", "INFY", "IOC", "ITC",
                      "JINDALSTEL", "JUBLFOOD", "JUSTDIAL", "KOTAKBANK", "L&TFH", "LICHSGFIN", "LT", "LUPIN", "M&M",
                      "M&MFIN", "MANAPPURAM", "MARICO", "MARUTI", "MCDOWELL-N", "MFSL", "MINDTREE", "MOTHERSUMI", "MRF",
                      "MUTHOOTFIN", "POWERGRID", "PVR", "NAUKRI", "NESTLEIND", "NIITTECH", "NMDC", "NTPC", "OIL", "ONGC",
                      "PAGEIND", "PEL", "PETRONET", "PFC", "PIDILITIND", "NIFTY 50", "NIFTY BANK", "SENSEX",
                      "NIFTY MIDCAP 100", "USDINR20MAYFUT", "INDIA VIX", "NIFTY20MAYFUT", "NIFTY20DEC9000CE",
                      "NIFTY20DEC9000PE", "RAMCOCEM", "RBLBANK", "RELIANCE", "SAIL", "RECLTD", "SBIN", "SHREECEM",
                      "SIEMENS", "SRF", "SRTRANSFIN", "SUNTV", "SUNPHARMA", "TATACHEM", "TATACONSUM", "TATAMOTORS",
                      "TATAPOWER", "TVSMOTOR", "UBL", "UJJIVAN", "ULTRACEMCO", "UPL", "VEDL", "VOLTAS", "WIPRO", "YESBANK", "ZEEL"]

    print("Sheet again opened")
    filename1 = os.path.join(os.path.dirname(__file__), 'Zerodha Instruments.xlsx')
    wb1 = load_workbook(filename1)
    ws1 = wb1.active
    print("Created new sheet")
    ws2 = wb1.create_sheet("Watchlist")

    i = 2
    sequences = []
    while i <= ws1.max_row:
        if ws1.cell(row=i, column=3).value in trading_symbol:
            sequences.append(i)
            print(i)
        i += 1

    columns = ws1.max_column
    i = 2
    print("Copying in new sheet")
    ws2.append(ordered_list)
    for row in sequences:
        col = 1
        while col <= columns:
            ws2.cell(row=i, column=col).value = ws1.cell(row=row, column=col).value
            col += 1
        i += 1

    print("Copied")
    wb1.save(filename)
    wb1.close()


def write_key_to_settings(key, value):
    filename = os.path.join(gettempdir(), 'zerodha_api.json')
    try:
        file = open(filename, 'r')
    except IOError:
        data = {"api_key": "", "api_secret": "", "redirect_uri": "", "access_token": ""}
        with open(filename, 'w') as output_file:
            json.dump(data, output_file)
    file = open(filename, 'r')
    try:
        data = json.load(file)
    except:
        data = {}
    data[key] = value
    with open(filename, 'w') as output_file:
        json.dump(data, output_file)


def read_key_from_settings(key):
    filename = os.path.join(gettempdir(), 'zerodha_api.json')
    try:
        file = open(filename, 'r')
    except IOError:
        file = open(filename, 'w')
    file = open(filename, 'r')
    try:
        data = json.load(file)
        return data[key]
    except:
        pass
    return None


def authenticate(login_url, username, password, pin):
    print("Inside authenticate")
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_exe = os.path.join(os.path.dirname(__file__), 'chromedriver.exe')
    driver = webdriver.Chrome(executable_path=chrome_exe, chrome_options=chrome_options)
    print("Login url inside authenticate: {0}".format(login_url))
    driver.get(login_url)
    WebDriverWait(driver, timeout=60).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/div/div[2]/div[1]/div/div/div[2]/form/div[1]/input")))
    print("Login site")
    driver.find_element_by_xpath('/html/body/div[1]/div/div[2]/div[1]/div/div/div[2]/form/div[1]/input').send_keys(username)
    print("Username: {0}".format(username))
    driver.find_element_by_xpath('/html/body/div[1]/div/div[2]/div[1]/div/div/div[2]/form/div[2]/input').send_keys(password)
    print("Password: {0}".format(password))
    driver.find_element_by_xpath('/html/body/div[1]/div/div[2]/div[1]/div/div/div[2]/form/div[4]/button').click()
    WebDriverWait(driver, timeout=60).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/div/div[2]/div[1]/div/div/div[2]/form/div[2]/div/input")))
    print("Page to enter Pin")
    driver.find_element_by_xpath('/html/body/div[1]/div/div[2]/div[1]/div/div/div[2]/form/div[2]/div/input').send_keys(pin)
    print("Pin: {0}".format(pin))
    driver.find_element_by_xpath('/html/body/div[1]/div/div[2]/div[1]/div/div/div[2]/form/div[3]/button').click()

    print("Current URL: {0}".format(driver.current_url))

    if not re.match("^.*.request_token=(.*)", driver.current_url):
        print("Approving User")
        print("Current URL: {0}".format(driver.current_url))
        try:
            WebDriverWait(driver, timeout=20).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/div/div[1]/div/div/div[3]/button")))
            driver.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/div/div[3]/button").click()
        except:
            pass
        print("Approved User")

    parsed = urlparse.urlparse(driver.current_url)
    access_code = parse_qs(parsed.query)['request_token'][0]
    print("Access code: {0}".format(access_code))
    return access_code


def initial_setup():
    global s, u, data, kws, loop

    try:
        logged_in = False

        stored_api_key = read_key_from_settings('api_key')
        stored_access_token = read_key_from_settings('access_token')
        if stored_access_token is not None and stored_api_key is not None:
            print('You already have a stored access token: [%s] paired with API key [%s]' % (
                stored_access_token, stored_api_key))

            try:
                u = KiteConnect(api_key=stored_api_key)
                u.set_access_token(stored_access_token)
                kws = KiteTicker(stored_api_key, stored_access_token)
                u.profile()
                logged_in = True
            except KiteException.TokenException as e:
                print('Sorry, there was an error [%s]. Let''s start over\n\n' % e)

        if logged_in is False:

            stored_api_key = read_key_from_settings('api_key')
            if stored_api_key is None:
                stored_api_key = input('What is your app''s API key [%s]:  ' % stored_api_key)
            write_key_to_settings('api_key', stored_api_key)

            stored_api_secret = read_key_from_settings('api_secret')
            if stored_api_secret is None:
                stored_api_secret = input('What is your app''s API secret [%s]:  ' % stored_api_secret)
            write_key_to_settings('api_secret', stored_api_secret)

            stored_redirect_uri = read_key_from_settings('redirect_uri')
            if stored_redirect_uri is None:
                stored_redirect_uri = input('What is your app''s redirect_uri [%s]:  ' % stored_redirect_uri)
            write_key_to_settings('redirect_uri', stored_redirect_uri)

            stored_username = read_key_from_settings('username')
            if stored_username is None:
                stored_username = input('What is your Zerodha username:  ')
            write_key_to_settings('username', stored_username)

            stored_password = read_key_from_settings('password')
            if stored_password is None:
                stored_password = input('What is your Zerodha password:  ')
            write_key_to_settings('password', stored_password)

            stored_password2fa = read_key_from_settings('password2fa')
            if stored_password2fa is None:
                stored_password2fa = input('What is your Zerodha Pin:  ')
            write_key_to_settings('password2fa', stored_password2fa )

            u = KiteConnect(api_key=stored_api_key)

            print('URL: %s\n' % u.login_url())

            try:
                print("Trying to authenticate")
                token = authenticate(u.login_url(), stored_username, stored_password, stored_password2fa)
                print("Token: {0}".format(token))
            except SystemError as se:
                print('Uh oh, there seems to be something wrong. Error: [%s]' % se)
                return

            print("Api secret: {0}".format(stored_api_secret))
            data = u.generate_session(token, api_secret=stored_api_secret)
            print("Data: {0}".format(data))
            write_key_to_settings('access_token', data['access_token'])
            u.set_access_token(data['access_token'])
            kws = KiteTicker(stored_api_key, data['access_token'])

        # kws = KiteTicker(stored_api_key, stored_access_token)
        # kws.on_ticks = on_ticks
        kws.on_connect = on_connect
        kws.on_close = on_close

        kws.connect(threaded=True)

        while True:
            def on_ticks(ws, ticks):
                print("Ticks= {}".format(ticks))
                helper_method(ticks)

            def helper_method(ticks):
                print("Inside helper function")
                filename = os.path.join(os.path.dirname(__file__), "Zerodha Instruments.xlsx")
                pythoncom.CoInitialize()
                app = xw.App(visible=False)
                wb = app.books.open(filename)

                ws = wb.sheets['Watchlist']
                df = pd.read_excel(filename, sheet_name='Watchlist')
                instrument_tokens_list = df['instrument_token'].tolist()
                for tick in ticks:
                    if tick['instrument_token'] in instrument_tokens_list:
                        pos = instrument_tokens_list.index(tick['instrument_token'])
                        cell = 'E' + str(pos + 2)
                        ws.range(cell).value = tick['last_price']

                        # Code to place an order
                        # transact_type_cell = 'N' + str(pos + 2)
                        # transact_type = ws.range(transact_type_cell).value
                        # trading_symbol_cell = 'C' + str(pos + 2)
                        # trading_symbol = ws.range(trading_symbol_cell).value
                        # quantity_cell = 'O' + str(pos + 2)
                        # quantity = ws.range(quantity_cell).value
                        #
                        # if transact_type == 'buy':
                        #     if trading_symbol == 'NSE':
                        #         u.place_order(tradingsymbol=trading_symbol, exchange=u.EXCHANGE_NSE,
                        #                       transaction_type=u.TRANSACTION_TYPE_BUY, quantity=quantity,
                        #                       order_type=u.ORDER_TYPE_MARKET, product=u.PRODUCT_NRML)
                        #
                        #     elif trading_symbol == 'BSE':
                        #         u.place_order(tradingsymbol=trading_symbol, exchange=u.EXCHANGE_BSE,
                        #                       transaction_type=u.TRANSACTION_TYPE_BUY, quantity=quantity,
                        #                       order_type=u.ORDER_TYPE_MARKET, product=u.PRODUCT_NRML)
                        #
                        #     elif trading_symbol == 'MCX':
                        #         u.place_order(tradingsymbol=trading_symbol, exchange=u.EXCHANGE_MCX,
                        #                       transaction_type=u.TRANSACTION_TYPE_BUY, quantity=quantity,
                        #                       order_type=u.ORDER_TYPE_MARKET, product=u.PRODUCT_NRML)
                        #
                        #     elif trading_symbol == 'NFO':
                        #         u.place_order(tradingsymbol=trading_symbol, exchange=u.EXCHANGE_NFO,
                        #                       transaction_type=u.TRANSACTION_TYPE_BUY, quantity=quantity,
                        #                       order_type=u.ORDER_TYPE_MARKET, product=u.PRODUCT_NRML)
                        #
                        # elif transact_type == 'sell':
                        #     if trading_symbol == 'NSE':
                        #         u.place_order(tradingsymbol=trading_symbol, exchange=u.EXCHANGE_NSE,
                        #                       transaction_type=u.TRANSACTION_TYPE_SELL, quantity=quantity,
                        #                       order_type=u.ORDER_TYPE_MARKET, product=u.PRODUCT_NRML)
                        #
                        #     elif trading_symbol == 'BSE':
                        #         u.place_order(tradingsymbol=trading_symbol, exchange=u.EXCHANGE_NSE,
                        #                       transaction_type=u.TRANSACTION_TYPE_SELL, quantity=quantity,
                        #                       order_type=u.ORDER_TYPE_MARKET, product=u.PRODUCT_NRML)
                        #
                        #     elif trading_symbol == 'MCX':
                        #         u.place_order(tradingsymbol=trading_symbol, exchange=u.EXCHANGE_MCX,
                        #                       transaction_type=u.TRANSACTION_TYPE_SELL, quantity=quantity,
                        #                       order_type=u.ORDER_TYPE_MARKET, product=u.PRODUCT_NRML)
                        #
                        #     elif trading_symbol == 'NFO':
                        #         u.place_order(tradingsymbol=trading_symbol, exchange=u.EXCHANGE_NFO,
                        #                       transaction_type=u.TRANSACTION_TYPE_SELL, quantity=quantity,
                        #                       order_type=u.ORDER_TYPE_MARKET, product=u.PRODUCT_NRML)
                        #
                        # ws.range(transact_type_cell).value = ""
                        # ws.range(quantity_cell).value = ""
                wb.save(filename)
                app.quit()

            kws.on_ticks = on_ticks
    except Exception as error:
        print("Error {0}".format(str(error)))
        exit(0)